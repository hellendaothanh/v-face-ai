import csv
import io
from datetime import datetime
from typing import Any, Dict, List, Optional
from loguru import logger

# openpyxl for Excel generation
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# reportlab for PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, inch, mm
from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


class ReportExportService:
    """
    Enterprise Report Generation & Export Service
    Supports Excel (.xlsx), PDF Document (.pdf), and UTF-8 BOM CSV (.csv).
    """

    # -------------------------------------------------------------------------
    # 1. EXCEL EXPORT (.xlsx)
    # -------------------------------------------------------------------------
    @staticmethod
    def build_attendance_excel(
        records: List[Dict[str, Any]],
        title: str = "BÁO CÁO DỮ LIỆU ĐIỂM DANH & CHẤM CÔNG DOANH NGHIỆP",
        date_range_str: str = ""
    ) -> io.BytesIO:
        """
        Creates a formatted .xlsx workbook with custom color palettes, borders,
        and auto-adjusted column widths.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dữ Liệu Chấm Công"
        ws.views.sheetView[0].showGridLines = True

        # Palettes
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Navy Blue
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        title_font = Font(name="Arial", size=15, bold=True, color="1E293B")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=9, color="334155")
        bold_font = Font(name="Arial", size=9, bold=True, color="1E293B")
        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1")
        )

        # Title Block
        ws.merge_cells("A1:J1")
        ws["A1"] = f"V-FACE AI ENTERPRISE • {title.upper()}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Subtitle / Date range
        ws.merge_cells("A2:J2")
        ws["A2"] = f"Thời gian trích xuất: {date_range_str or datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | Tổng số bản ghi: {len(records)}"
        ws["A2"].font = Font(name="Arial", size=9, italic=True, color="64748B")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18

        ws.append([])  # Row 3 blank

        # Table Headers
        headers = [
            "STT",
            "Mã Nhân Viên",
            "Họ và Tên",
            "Phòng Ban",
            "Thời Gian",
            "Hình Thức",
            "Độ Khớp Biometrics",
            "Đồ Bảo Hộ PPE",
            "Ca Áp Dụng",
            "Ghi Chú"
        ]
        ws.append(headers)
        header_row = 4
        ws.row_dimensions[header_row].height = 24

        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Populate Data Rows
        for i, r in enumerate(records, start=1):
            row_idx = header_row + i
            ws.row_dimensions[row_idx].height = 20

            # Formatted Values
            check_time_str = r.get("check_time")
            if isinstance(check_time_str, datetime):
                check_time_str = check_time_str.strftime("%d/%m/%Y %H:%M:%S")

            conf_val = f"{float(r.get('confidence_score') or 0):.1f}%"
            ppe_str = "Tuân thủ (Đầy đủ)" if r.get("ppe_compliance", True) else f"Vi phạm ({r.get('ppe_violations', 'Thiếu PPE')})"
            
            row_data = [
                i,
                r.get("employee_code", "N/A"),
                r.get("employee_name", "N/A"),
                r.get("department", "N/A"),
                check_time_str or "",
                r.get("attendance_type", "CHECK_IN"),
                conf_val,
                ppe_str,
                r.get("shift_code", "SHIFT_STANDARD"),
                r.get("note", "") or ""
            ]
            ws.append(row_data)

            # Apply cell styles
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.font = data_font
                c.border = thin_border
                if i % 2 == 0:
                    c.fill = zebra_fill
                
                # Alignments
                if col_idx in [1, 2, 5, 6, 7, 9]:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-fit Column Widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len and cell.row > 2:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # -------------------------------------------------------------------------
    # 2. PDF EXPORT (.pdf)
    # -------------------------------------------------------------------------
    @staticmethod
    def build_attendance_pdf(
        records: List[Dict[str, Any]],
        title: str = "BÁO CÁO CHẤM CÔNG & SINH TRẮC HỌC FACE AI",
        date_range_str: str = ""
    ) -> io.BytesIO:
        """
        Creates an executive PDF Report document using reportlab Platypus.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm
        )

        elements = []
        styles = getSampleStyleSheet()

        # Custom Styles
        title_style = ParagraphStyle(
            name="DocTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            textColor=colors.HexColor("#1E3A8A"),
            alignment=1  # Center
        )
        subtitle_style = ParagraphStyle(
            name="DocSubTitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#64748B"),
            alignment=1
        )
        cell_style = ParagraphStyle(
            name="CellText",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#334155")
        )
        header_cell_style = ParagraphStyle(
            name="HeaderCell",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
            alignment=1
        )

        # Header Block
        elements.append(Paragraph("V-FACE AI BIOMETRICS & HRM ENTERPRISE", title_style))
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 4))
        meta_text = f"Thời gian xuất: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | Phạm vi: {date_range_str or 'Toàn thời gian'} | Tổng số bản ghi: {len(records)}"
        elements.append(Paragraph(meta_text, subtitle_style))
        elements.append(Spacer(1, 12))
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CBD5E1"), spaceAfter=12))

        # Table Data
        table_headers = [
            Paragraph("STT", header_cell_style),
            Paragraph("Mã NV", header_cell_style),
            Paragraph("Họ và Tên", header_cell_style),
            Paragraph("Phòng Ban", header_cell_style),
            Paragraph("Thời Gian", header_cell_style),
            Paragraph("Loại", header_cell_style),
            Paragraph("Độ Khớp", header_cell_style),
            Paragraph("Đồ Bảo Hộ PPE", header_cell_style),
            Paragraph("Ca Làm Việc", header_cell_style),
        ]
        table_data = [table_headers]

        for i, r in enumerate(records[:150], start=1):  # Cap to first 150 rows for clean PDF page limits
            check_time_str = r.get("check_time")
            if isinstance(check_time_str, datetime):
                check_time_str = check_time_str.strftime("%d/%m/%Y %H:%M")

            conf_str = f"{float(r.get('confidence_score') or 0):.1f}%"
            ppe_status = "Đạt" if r.get("ppe_compliance", True) else "Vi phạm"

            row = [
                Paragraph(str(i), cell_style),
                Paragraph(r.get("employee_code", "N/A"), cell_style),
                Paragraph(r.get("employee_name", "N/A"), cell_style),
                Paragraph(r.get("department", "N/A"), cell_style),
                Paragraph(check_time_str or "", cell_style),
                Paragraph(str(r.get("attendance_type", "CHECK_IN")), cell_style),
                Paragraph(conf_str, cell_style),
                Paragraph(ppe_status, cell_style),
                Paragraph(r.get("shift_code", "STANDARD"), cell_style),
            ]
            table_data.append(row)

        # Build Table
        t = Table(table_data, colWidths=[1.2 * cm, 2.5 * cm, 4.5 * cm, 3.5 * cm, 3.5 * cm, 2.5 * cm, 2.2 * cm, 2.8 * cm, 3.0 * cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)

        # Footer
        elements.append(Spacer(1, 15))
        footer_p = Paragraph("<i>Báo cáo được khởi tạo tự động bởi Hệ thống Quản trị Nhân sự & Sinh trắc học V-Face AI.</i>", subtitle_style)
        elements.append(footer_p)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    # -------------------------------------------------------------------------
    # 3. CSV EXPORT (.csv)
    # -------------------------------------------------------------------------
    @staticmethod
    def build_attendance_csv(records: List[Dict[str, Any]]) -> str:
        """
        Creates a UTF-8 BOM CSV string compatible with Excel, Sheets, and analytics tools.
        """
        output = io.StringIO()
        writer = csv.writer(output, dialect="excel")

        headers = [
            "STT",
            "Mã Nhân Viên",
            "Họ và Tên",
            "Phòng Ban",
            "Thời Gian",
            "Loại Điểm Danh",
            "Tỷ Lệ Khớp Face AI (%)",
            "Đồ Bảo Hộ PPE",
            "Chi Tiết Vi Phạm PPE",
            "Mã Ca Làm Việc",
            "Ghi Chú"
        ]
        writer.writerow(headers)

        for i, r in enumerate(records, start=1):
            check_time_str = r.get("check_time")
            if isinstance(check_time_str, datetime):
                check_time_str = check_time_str.strftime("%Y-%m-%d %H:%M:%S")

            writer.writerow([
                i,
                r.get("employee_code", ""),
                r.get("employee_name", ""),
                r.get("department", ""),
                check_time_str or "",
                r.get("attendance_type", "CHECK_IN"),
                f"{float(r.get('confidence_score') or 0):.2f}",
                "Tuân thủ" if r.get("ppe_compliance", True) else "Vi phạm",
                r.get("ppe_violations", "") or "",
                r.get("shift_code", "SHIFT_STANDARD"),
                r.get("note", "") or ""
            ])

        return "\ufeff" + output.getvalue()
